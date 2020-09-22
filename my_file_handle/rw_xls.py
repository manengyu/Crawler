#coding:utf-8
import urllib
import urllib2
import requests
import re
import time
import xlrd
import xlwt
import sys
reload(sys)
sys.setdefaultencoding("utf-8")


def read_xls(sheet, file_path, *args):
    workbook = xlrd.open_workbook(file_path)
    booksheet = workbook.sheet_by_name(sheet)
    ls = []
    for row in range(booksheet.nrows):
        if booksheet.cell(row, 4).value == args[1]:
            if args[0] == "date":
                tuple_date = xlrd.xldate_as_tuple(booksheet.cell(row, 5).value, 0)  # 读取日期时
                str_date = str(tuple_date[0]).decode("utf-8") + u"/" + str(tuple_date[1]).decode("utf-8") + u"/" +\
                           str(tuple_date[2]).decode("utf-8")
                if str_date not in ls:
                    ls.append(str_date)
                ls.append(booksheet.cell(row, 1).value)
                ls.append(booksheet.cell(row, 2).value)
                ls.append(booksheet.cell(row, 3).value)
    return ls

def init_xls():
    new_workbook = xlwt.Workbook()
    sheet_top20 = new_workbook.add_sheet(u"")
    sheet_all = new_workbook.add_sheet(u"")
    sheet_api = new_workbook.add_sheet(u"")
    sheet_amount = new_workbook.add_sheet(u"")
    sheet_amount_7 = new_workbook.add_sheet(u"")
    sheet_error_url = new_workbook.add_sheet(u"")
    title_amount = [u"", u"", u"", u""]
    title_api = [u"", u"", u"", u"", u"", u""]
    title_data = [u"", u"", u"", u""]
    title_errorurl = [u"", u"", u""]
    for j in range(0, len(title_amount)):
        sheet_amount.write(0, j, title_amount[j])
        sheet_amount_7.write(0, j, title_amount[j])
    for j in range(0, len(title_api)):
        sheet_api.write(0, j, title_api[j])
    for j in range(0, len(title_data)):
        sheet_top20.write(0, j, title_data[j])
        sheet_all.write(0, j, title_data[j])
    for j in range(0, len(title_errorurl)):
        sheet_error_url.write(0, j, title_errorurl[j])
    return new_workbook, sheet_amount, sheet_amount_7, sheet_api, sheet_top20, sheet_all, sheet_error_url


def init_xls(title=None, sheet=None, new_workbook=None):
    if not new_workbook:
        new_workbook = xlwt.Workbook()
    sheet = new_workbook.add_sheet(sheet or u"sheet1")
    for jj in range(0, len(title)):
        sheet.write(0, jj, title[jj])
    return new_workbook, sheet
    
    
def write_xls(sheet, nu,  *args):
    for i in range(len(args)):
        sheet.write(nu+1, i, args[i])  # 存在标题，所以+1


def save_xls(file_name, new_workbook):
    new_workbook.save(file_name)


def main(sheet, file_path, user_name):  # 表格左上角为（0，0），文件名及内容统一用unicode
    ls_name = read_xls(sheet, file_path, u"user_name", user_name)
    new_workbook, sheet_amount, sheet_amount_7, sheet_api, sheet_top20, sheet_all, sheet_error_url = init_xls()
    save_xls(file_path, new_workbook)

# openpyxl==2.6.2
def flask_xls():
    with open("./batch.xls", 'wb') as f:
        f.write(data)
    wb = xlrd.open_workbook("./batch.xls")
    sheet1 = wb.sheet_by_index(0)
    rowNum = sheet1.nrows
    batchtolist = list()

    title = [sheet1.row_values(0)[0], sheet1.row_values(0)[1], sheet1.row_values(0)[2], "是否一致"]
    for i in range(rowNum - 1):
        batchtolist.append([sheet1.row_values(i + 1)[0], sheet1.row_values(i + 1)[1], sheet1.row_values(i + 1)[2]])

    for i in batchtolist:
        response = get_data("identity", rm_space(i[0]), rm_space(i[1]), None, None, None, None, None, None,
                                    source)
        i.append(json.loads(response.text)["isMatch"] if response.status_code == 200 else "无法核查")

    new_workbook, sheet = init_xls(title=title)
    for n, i in enumerate(batchtolist):
        write_xls(sheet, n, i)
        
    bio = BytesIO()
    new_workbook.save(bio)
    bio.seek(0)
    return bio

    import openpyxl
    from io import BytesIO
    from flask import make_response
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    fields = ["", "", "", "", "", "", "", "", "", "", "创建时间", "", "", ""]
    for f in range(0, len(fields)):
        sheet.cell(1, f + 1, fields[f])
    row_offset = 1

    for row, i in enumerate(resList):
        sheet.cell(row_offset + row + 1, 1, str(i["_id"]))
        sheet.cell(row_offset + row + 1, 1).hyperlink = f""
        sheet.cell(row_offset + row + 1, 2, i])
        sheet.cell(row_offset + row + 1, 3, i)
        sheet.cell(row_offset + row + 1, 4, str(time.strftime("%Y-%m-%d", time.localtime(i.get("dueDate"))) if i.get("dueDate") else ""))

    bio = BytesIO()
    workbook.save(bio)
    bio.seek(0)
    response = make_response(bio.getvalue())
    bio.close()
    response.headers['Content-Type'] = 'application/vnd.ms-excel'  # 文件类型
    response.headers['Content-Disposition'] = f"attachment;filename*=UTF-8''{time.strftime('%Y%m%d', time.localtime(time.time()))}.xlsx"
    return response

# 测试函数
if __name__ == u"__main__":
    main(u'Sheet1', ur'D:\cncepgf\workspace\07.07-07.28.xlsx', u"")  # 工作表 文件所在路径 名字




